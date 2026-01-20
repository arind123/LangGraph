import io

import re

import pandas as pd

import openpyxl

import numpy as np

import pytest

from openpyxl.utils import get_column_letter

from typing import Dict, List, Any, Optional

from pydantic import BaseModel

from langgraph.graph import StateGraph, END, START

from langchain_openai import ChatOpenAI

from contextlib import redirect_stdout

from dotenv import load_dotenv

import traceback



load_dotenv()



# --- STEP 1: DUAL-CONTEXT MINER ---

class IndustryLogicMiner:

    def __init__(self, file_path: str):

        self.file_path = file_path

        self.wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)

        self.val_wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)



    def _get_header_map(self, ws_v) -> Dict[str, str]:

        header_row = next(ws_v.iter_rows(min_row=1, max_row=1), [])

        return {get_column_letter(i): re.sub(r'[^a-zA-Z0-9_]', '_', str(c.value or f"Col_{i}"))

                for i, c in enumerate(header_row, 1)}



    def _map_formula(self, formula: str, header_map: Dict[str, str]) -> Dict[str, str]:

        if not isinstance(formula, str): return {"raw": "", "semantic": ""}

        row_match = re.search(r'\d+', formula)

        row_num = row_match.group(0) if row_match else ""

        raw_pattern = formula.replace(row_num, "{n}") if row_num else formula

       

        def to_header(match):

            col = match.group(1)

            return f"df['{header_map.get(col, col)}']"

       

        semantic_pattern = re.sub(r'\$?([A-Z]+)\$?\d+', to_header, formula)

        return {"excel_pattern": raw_pattern, "python_semantic": semantic_pattern}



    def extract_full_context(self) -> Dict[str, Any]:

        model_meta = {}

        for sheet_name in self.wb.sheetnames:

            # ws has formulas, ws_v has values

            ws = self.wb[sheet_name]

            ws_v = self.val_wb[sheet_name]

           

            header_map = self._get_header_map(ws_v)

            vector_rules = {}

           

            # 1. EXTRACT LOGIC FROM 'ws' (The one WITH formulas)

            # Use ws here, not ws_v

            for row in ws.iter_rows(min_row=2, max_row=11):

                for col_idx, cell in enumerate(row, start=1):

                    # In ws, cell.value will be the string "=VLOOKUP..."

                    if cell and isinstance(cell.value, str) and cell.value.startswith('='):

                        col_letter = get_column_letter(col_idx)

                        header_name = header_map[col_letter]

                       

                        if header_name not in vector_rules:

                            patterns = self._map_formula(cell.value, header_map)

                            vector_rules[header_name] = {

                                "excel_col": col_letter,

                                **patterns

                            }

           

            # 2. EXTRACT SCHEMA FROM 'ws_v' (The one WITH calculated values)

            df_sample = pd.DataFrame([

                {header_map[get_column_letter(i)]: c.value for i, c in enumerate(r, 1)}

                for r in ws_v.iter_rows(min_row=2, max_row=100)

            ]).dropna(how='all')



            schema = {}

            # ... (Rest of your schema logic remains the same) ...

            for col in df_sample.columns:

                series = df_sample[col]

                if pd.api.types.is_numeric_dtype(series):

                    # Ensure we handle empty sheets to avoid min/max errors

                    s_min = series.min()

                    s_max = series.max()

                    schema[col] = {

                        "type": "numeric",

                        "min": float(s_min) if pd.notnull(s_min) else 0,

                        "max": float(s_max) if pd.notnull(s_max) else 0

                    }

                else:

                    schema[col] = {

                        "type": "categorical",

                    }

           

            model_meta[sheet_name] = {"logic": vector_rules, "schema": schema}

           

        return model_meta

       



# --- STEP 2: AGENT STATE ---

class AgentState(BaseModel):

    excel_path: str

    metadata: Dict[str, Any] = {}

    generated_code: str = ""

    unit_tests: str = ""

    test_results: str = ""

    error_log: Optional[str] = None

    iterations: int = 0

    is_validated: bool = False



# --- STEP 3: NODES ---

llm = ChatOpenAI(model="gpt-4o", temperature=0)



def miner_node(state: AgentState):

    miner = IndustryLogicMiner(state.excel_path)

    print(f"Excel Extract: {miner.extract_full_context()}\n")

    return {"metadata": miner.extract_full_context()}



def architect_node(state: AgentState):

    feedback = f"\nFIX ERROR FROM PREVIOUS ATTEMPT: {state.error_log}" if state.error_log else ""

    prompt = f"""

    You are a Senior Python Developer. Translate Excel sheet logic into a vectorized Python class.

   

    ENVIRONMENT CONTEXT:

    - You have access to a dictionary 'all_data' where keys are sheet names and values are DataFrames.

    - The active sheet is passed as 'df'.

   

    METADATA, FORMULAS & SCHEMA CONTEXT (No real data):

    {state.metadata}

   

    {feedback}

   

    REQUIREMENTS:

    1. Define 'ExcelCalculator(df: pd.DataFrame, all_data: Dict[str, pd.DataFrame])'.

    2. Map every 'logic' from the metadata to a @property.

    3. If a formula references another sheet (e.g., 'Products!A1'), access it via 'self.all_data["Products"]'.

    4. Use purely vectorized Pandas/NumPy operations. Do not hardcode values from samples.

    5. Return ONLY the code inside a markdown code block. No need to be chatty.

    6. Use the 'schema' to determine if columns are floats, ints, or strings.

    7. Ensure code handles potential NaNs if ranges suggest empty cells.

    """

    res = llm.invoke(prompt)

    code_match = re.search(r"```python\s+(.*?)\s+```", res.content, re.DOTALL)

    if code_match:

        code = code_match.group(1)

    else:

        # Fallback if no backticks

        code = res.content

    print(f"HumanMessage\n: {prompt} \n AIMessage\n: {res.content} \n OutpuCode\n: {code}")

    return {"generated_code": code, "iterations": state.iterations + 1}





def sandbox_node(state: AgentState):

    print(f"\n--- 🔒 ANONYMOUS STRUCTURE TEST ---")

    try:

        exec_context = {"pd": pd, "np": np, "Dict": Dict, "Any": Any}

        exec(state.generated_code, exec_context)

        CalculatorClass = exec_context.get("ExcelCalculator")



        # GENERATE FULLY ANONYMOUS TEST DATA

        synthetic_context = {}

        for sheet, meta in state.metadata.items():

            rows = 10

            dummy_data = {}

            for col, info in meta['schema'].items():

                if info['type'] == 'numeric':

                    # Create range-appropriate floats

                    dummy_data[col] = np.random.uniform(float(info['min']), float(info['max']), rows).astype(float)

                else:

                    # Create generic placeholders like "ID_0", "ID_1"

                    dummy_data[col] = [f"PLACEHOLDER_{i}" for i in range(rows)]

            synthetic_context[sheet] = pd.DataFrame(dummy_data)



        # print(synthetic_context)



        # TEST EXECUTION

        for sheet, meta in state.metadata.items():

            if not meta['logic']: continue

            calc = CalculatorClass(df=synthetic_context[sheet], all_data=synthetic_context)

           

            for col in meta['logic'].keys():

                actual = getattr(calc, col.lower())

                # If it executes without error, the logic is structurally sound

                print(f" ✅ Structural Match: {sheet} -> {col}")



        return {"is_validated": True, "error_log": None}

    except Exception as e:

        return {"is_validated": False, "error_log": f"Structure Error: {str(e)}"}

   





def test_gen_node(state: AgentState):

    prompt = f"""
    Write a pytest file for this code:
    {state.generated_code}
    
    Metadata: {state.metadata}
    
    REQUIREMENTS:
    1. Import the calculator as 'from model import ExcelCalculator'.
    2. Create a demo dataset where ALL numeric columns are strictly float type.
    3. IMPORTANT: For every test assertion, include a message that prints the expected value and the actual value.
       Example: assert actual == expected, f"Failed! Expected: {{expected}}, but got: {{actual}}"
    4. Ensure the tests cover cross-sheet logic.
    """

    res = llm.invoke(prompt)

    code_match = re.search(r"```python\s+(.*?)\s+```", res.content, re.DOTALL)

    if code_match:

        code = code_match.group(1)

    else:

        # Fallback if no backticks

        code = res.content



    return {"unit_tests": code}



def execute_tests_node(state: AgentState):

    with open("model.py", "w") as f: f.write(state.generated_code)

    with open("test_model.py", "w") as f: f.write(state.unit_tests)

    # 3. Create the demo XLSX with float precision
    output_xlsx = "pytest_demo_data.xlsx"
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        for sheet, meta in state.metadata.items():
            rows = 10
            dummy_data = {}
            for col, info in meta['schema'].items():
                if info['type'] == 'numeric':
                    # Explicitly casting to float
                    dummy_data[col] = np.random.uniform(float(info['min']), float(info['max']), rows).astype(float)
                else:
                    dummy_data[col] = [f"PLACEHOLDER_{i}" for i in range(rows)]
            pd.DataFrame(dummy_data).to_excel(writer, sheet_name=sheet, index=False)

    f_io = io.StringIO()

    with redirect_stdout(f_io):

        pytest.main(["-v", "-s", "test_model.py"])

    return {"test_results": f_io.getvalue()}





# --- STEP 4: GRAPH ASSEMBLY ---

workflow = StateGraph(AgentState)

workflow.add_node("mine", miner_node)

workflow.add_node("architect", architect_node)

workflow.add_node("sandbox", sandbox_node)

workflow.add_node("test_gen", test_gen_node)

workflow.add_node("run_tests", execute_tests_node)



workflow.add_edge(START, "mine")

workflow.add_edge("mine", "architect")

workflow.add_edge("architect", "sandbox")



def router(state: AgentState):

    if state.is_validated: return "generate"

    return "retry" if state.iterations < 5 else "end"



workflow.add_conditional_edges("sandbox", router, {"generate": "test_gen", "retry": "architect", "end": END})

workflow.add_edge("test_gen", "run_tests")

workflow.add_edge("run_tests", END)



agent = workflow.compile()





if __name__ == "__main__":

    final_result = agent.invoke({"excel_path": "complex_financial_model_4.xlsx"})

    print("\n--- PYTEST RESULTS ---\n", final_result.get("test_results", "No tests run."))